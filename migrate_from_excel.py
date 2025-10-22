import os
import sys
from datetime import datetime
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook


def find_onedrive_base() -> Optional[str]:
    for env in ("OneDriveCommercial", "OneDrive", "OneDriveConsumer"):
        p = os.environ.get(env)
        if p and os.path.isdir(p):
            return p
    home = os.path.expanduser("~")
    try:
        for name in os.listdir(home):
            if name.lower().startswith("onedrive"):
                p = os.path.join(home, name)
                if os.path.isdir(p):
                    return p
    except Exception:
        pass
    return None


def header_map(ws):
    return [(c.value if c is not None else None) for c in ws[1]]


def row_to_dict(headers, row):
    out = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        v = row[i].value if i < len(row) else None
        out[str(h)] = v if v is not None else ""
    return out


def locate_excel_path(arg_path: Optional[str]) -> Optional[str]:
    if arg_path:
        return arg_path if os.path.exists(arg_path) else None
    # 1) .env EXCEL_RELATIVE_PATH + OneDrive
    rel = os.getenv("EXCEL_RELATIVE_PATH")
    if rel:
        base = find_onedrive_base()
        if base:
            p = os.path.join(base, rel)
            if os.path.exists(p):
                return p
    # 2) Caminho padrão conhecido
    base = find_onedrive_base()
    if base:
        p = os.path.join(base, "01 LEANDRO", "IMPRESSÕES", "BANCO_DE_DADOS_ORCAMENTO.xlsx")
        if os.path.exists(p):
            return p
    # 3) Arquivo no diretório atual
    p = os.path.join(os.getcwd(), "BANCO_DE_DADOS_ORCAMENTO.xlsx")
    if os.path.exists(p):
        return p
    return None


def main():
    print("== Migração Excel -> DB ==")
    load_dotenv()

    try:
        from db_backend import DB
    except Exception as ex:
        print("ERRO: não foi possível importar db_backend:", ex)
        sys.exit(1)

    if not DB.is_ready():
        print("ERRO: DATABASE_URL inválido ou SQLAlchemy indisponível. Verifique seu .env.")
        sys.exit(2)

    # Garante esquema
    try:
        if hasattr(DB, "init_schema_portable"):
            DB.init_schema_portable()
        else:
            DB.init_schema()
    except Exception as ex:
        print("Aviso: falha ao criar esquema:", ex)

    arg_path = sys.argv[1] if len(sys.argv) > 1 else None
    excel_path = locate_excel_path(arg_path)
    if not excel_path:
        print("ERRO: não encontrei a planilha. Informe o caminho como argumento ou configure EXCEL_RELATIVE_PATH no .env.")
        sys.exit(3)

    print("Lendo:", excel_path)
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    def get_ws(*names):
        for n in names:
            if n in wb.sheetnames:
                return wb[n]
        return None

    ws_orc = get_ws("Orçamentos", "Orcamentos")
    ws_cad = get_ws("Cadastros")
    ws_ped = get_ws("Pedidos")

    from_count = {"orc": 0, "cad": 0, "ped": 0}
    to_count = {"orc": 0, "cad": 0, "ped": 0}

    if ws_orc is not None:
        headers = header_map(ws_orc)
        for i, row in enumerate(ws_orc.iter_rows(min_row=2), start=2):
            d = row_to_dict(headers, row)
            if not any(v for v in d.values()):
                continue
            from_count["orc"] += 1
            try:
                DB.salvar_orcamento(d)
                to_count["orc"] += 1
            except Exception as ex:
                print(f"  Linha {i} (Orçamentos) falhou: {ex}")

    if ws_cad is not None:
        headers = header_map(ws_cad)
        for i, row in enumerate(ws_cad.iter_rows(min_row=2), start=2):
            d = row_to_dict(headers, row)
            if not any(v for v in d.values()):
                continue
            from_count["cad"] += 1
            try:
                DB.salvar_cadastro(d)
                to_count["cad"] += 1
            except Exception as ex:
                print(f"  Linha {i} (Cadastros) falhou: {ex}")

    if ws_ped is not None:
        headers = header_map(ws_ped)
        for i, row in enumerate(ws_ped.iter_rows(min_row=2), start=2):
            d = row_to_dict(headers, row)
            if not any(v for v in d.values()):
                continue
            from_count["ped"] += 1
            try:
                DB.salvar_pedido(d)
                to_count["ped"] += 1
            except Exception as ex:
                print(f"  Linha {i} (Pedidos) falhou: {ex}")

    print("Concluído:")
    print("  Orçamentos:", from_count["orc"], "lidos /", to_count["orc"], "inseridos")
    print("  Cadastros :", from_count["cad"], "lidos /", to_count["cad"], "inseridos")
    print("  Pedidos   :", from_count["ped"], "lidos /", to_count["ped"], "inseridos")


if __name__ == "__main__":
    main()

